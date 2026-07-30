from flask import Flask, request, render_template, send_file
import openpyxl
import statistics
import io

app = Flask(__name__)

BAND_MAP = {"theta": 1, "gamma": 11, "high-gamma": 21}
DAY_MAP = {1: 3, 2: 11, 3: 19, 4: 27, 5: 35}
RATS = ["rat1", "rat2", "rat3"]

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/process", methods=["POST"])
def process():
    files = request.files.getlist("data_files")
    tpl = request.files.get("template_file")
    if not files or not tpl:
        return "请上传数据文件和模板", 400

    wb = openpyxl.load_workbook(tpl)
    ws = wb["Sheet1"]
    log = []

    for f in files:
        try:
            wd = openpyxl.load_workbook(f)
            sd = wd["Sheet1"]
            rat = sd.cell(2, 3).value
            day_str = sd.cell(2, 4).value
            band = sd.cell(2, 5).value
            day_num = int("".join(filter(str.isdigit, str(day_str))))

            start_row = BAND_MAP.get(band)
            if start_row is None:
                log.append(f"✗ {f.filename}: 未知波段")
                wd.close()
                continue
            if rat not in RATS:
                log.append(f"✗ {f.filename}: 未知rat")
                wd.close()
                continue
            col = DAY_MAP.get(day_num)
            if col is None:
                log.append(f"✗ {f.filename}: 未知day")
                wd.close()
                continue

            ri = RATS.index(rat)
            row_mean = start_row + 2 + ri * 2
            row_out = start_row + 3 + ri * 2

            for ch in range(1, 9):
                vals = []
                for r in range(2, 47):
                    v = sd.cell(r, 5 + ch).value
                    if v is not None:
                        vals.append(float(v))
                if vals:
                    m = statistics.mean(vals)
                    ws.cell(row_mean, col + ch - 1, value=m)
                    sv = sorted(vals)
                    q1 = sv[len(sv) // 4]
                    q3 = sv[3 * len(sv) // 4]
                    mid = [v for v in vals if q1 <= v <= q3]
                    fm = statistics.mean(mid) if mid else 0
                    ws.cell(row_out, col + ch - 1, value=fm)
            wd.close()
            log.append(f"✓ {f.filename}  ({rat}, {day_str}, {band})")
        except Exception as e:
            log.append(f"✗ {f.filename}: {str(e)}")

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name="tongji_processed.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
